"""Excel export with formulas, formatting, and multiple sheets."""

import io
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo


# Color palette
CLR_HEADER_BG = "1E3A5F"
CLR_HEADER_FG = "FFFFFF"
CLR_ANOMALY_BG = "FFE5E5"
CLR_ANOMALY_FG = "CC0000"
CLR_OK_BG = "E8F5E9"
CLR_OK_FG = "2E7D32"
CLR_ALT_BG = "F8F9FA"
CLR_ACCENT = "2563EB"


def _header_style(cell):
    cell.font = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", start_color=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="CCCCCC"),
    )


def _auto_width(ws, min_w=8, max_w=45):
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 3, max_w))


def _thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_xlsx(
    df_result: pd.DataFrame,
    threshold_pct: float,
    file_name: str = "analiza",
) -> bytes:
    """Generate full Excel report as bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_dane_sheet(wb, df_result)
    _build_anomalie_sheet(wb, df_result)
    _build_mediany_sheet(wb, df_result)
    _build_settings_sheet(wb, threshold_pct, file_name, df_result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_dane_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Dane")

    COLS = [
        "Index materiałowy", "Partia", "Nazwa materiału", "Magazyn",
        "Przyjęcie [PZ]", "Data przyjęcia", "Stan mag.", "Wartość mag.",
        "Cena", "Mediana ceny indeksu", "Cena referencyjna",
        "Źródło ceny referencyjnej", "Odchylenie % do mediany", "Anomalia",
    ]
    available = [c for c in COLS if c in df.columns]

    # Header row
    for col_idx, col_name in enumerate(available, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _header_style(cell)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Map column positions for formula references
    col_pos = {name: idx for idx, name in enumerate(available, 1)}
    stan_col = get_column_letter(col_pos.get("Stan mag.", 7))
    war_col = get_column_letter(col_pos.get("Wartość mag.", 8))
    cena_col = get_column_letter(col_pos.get("Cena", 9))
    ref_col = get_column_letter(col_pos.get("Cena referencyjna", 11))
    dev_col = get_column_letter(col_pos.get("Odchylenie % do mediany", 13))

    # Data rows
    for row_idx, (_, row) in enumerate(df[available].iterrows(), 2):
        for col_idx, col_name in enumerate(available, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name == "Cena":
                # Excel formula
                cell.value = f"={war_col}{row_idx}/{stan_col}{row_idx}"
                cell.number_format = "#,##0.0000"
            elif col_name == "Odchylenie % do mediany":
                cell.value = f"=IF({ref_col}{row_idx}<>0,ABS({cena_col}{row_idx}-{ref_col}{row_idx})/{ref_col}{row_idx}*100,\"\")"
                cell.number_format = "#,##0.00"
            elif col_name == "Anomalia":
                cell.value = f"=IF({dev_col}{row_idx}>=\"\",IF({dev_col}{row_idx}>={_get_threshold_from_settings(df)},\"TAK\",\"NIE\"),\"NIE\")"
            elif col_name == "Data przyjęcia":
                if pd.notna(val):
                    cell.value = val.date() if hasattr(val, "date") else val
                    cell.number_format = "YYYY-MM-DD"
            elif col_name in ("Stan mag.", "Wartość mag.", "Mediana ceny indeksu", "Cena referencyjna"):
                cell.value = float(val) if pd.notna(val) else None
                if col_name in ("Wartość mag.", "Mediana ceny indeksu", "Cena referencyjna"):
                    cell.number_format = "#,##0.00"
            else:
                cell.value = str(val) if pd.notna(val) else None

            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color=CLR_ALT_BG)
            cell.border = _thin_border()
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")

    # Conditional formatting: highlight anomaly rows
    anom_col_idx = col_pos.get("Anomalia")
    if anom_col_idx:
        anom_letter = get_column_letter(anom_col_idx)
        last_row = len(df) + 1
        data_range = f"A2:{get_column_letter(len(available))}{last_row}"
        red_fill = PatternFill("solid", start_color="FFE5E5")
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'${anom_letter}2="TAK"'],
                fill=red_fill,
                font=Font(color=CLR_ANOMALY_FG, name="Arial", size=9),
            ),
        )

    ws.auto_filter.ref = f"A1:{get_column_letter(len(available))}1"
    _auto_width(ws)


def _get_threshold_from_settings(df: pd.DataFrame) -> float:
    """Extract threshold - fallback to 20."""
    return 20


def _build_anomalie_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Anomalie")

    anomalies = df[df["Anomalia"] == True].copy()
    if anomalies.empty:
        ws["A1"] = "Brak anomalii cenowych"
        return

    COLS = [
        "Index materiałowy", "Partia", "Nazwa materiału", "Magazyn",
        "Przyjęcie [PZ]", "Data przyjęcia", "Stan mag.", "Wartość mag.",
        "Cena", "Cena referencyjna", "Źródło ceny referencyjnej",
        "Odchylenie % do mediany", "Anomalia",
    ]
    available = [c for c in COLS if c in anomalies.columns]

    for col_idx, col_name in enumerate(available, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _header_style(cell)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for row_idx, (_, row) in enumerate(anomalies[available].iterrows(), 2):
        for col_idx, col_name in enumerate(available, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == "Data przyjęcia":
                if pd.notna(val):
                    cell.value = val.date() if hasattr(val, "date") else val
                    cell.number_format = "YYYY-MM-DD"
            elif col_name in ("Stan mag.", "Wartość mag.", "Cena", "Cena referencyjna", "Mediana ceny indeksu"):
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = "#,##0.0000"
            elif col_name == "Odchylenie % do mediany":
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = "#,##0.00"
                cell.font = Font(name="Arial", bold=True, color=CLR_ANOMALY_FG, size=9)
            else:
                cell.value = str(val) if pd.notna(val) else None

            if col_name != "Odchylenie % do mediany":
                cell.fill = PatternFill("solid", start_color="FFF3F3")
                cell.font = Font(name="Arial", size=9)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(available))}1"
    _auto_width(ws)


def _build_mediany_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Mediany")

    medians_df = (
        df.groupby("Index materiałowy")
        .agg(
            Mediana=("Mediana ceny indeksu", "first"),
            Liczba_partii=("Partia", "count"),
            Liczba_anomalii=("Anomalia", "sum"),
        )
        .reset_index()
        .rename(columns={
            "Mediana": "Mediana ceny",
            "Liczba_partii": "Liczba partii",
            "Liczba_anomalii": "Liczba anomalii",
        })
        .sort_values("Liczba anomalii", ascending=False)
    )

    headers = list(medians_df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _header_style(cell)

    ws.freeze_panes = "A2"

    for row_idx, (_, row) in enumerate(medians_df.iterrows(), 2):
        for col_idx, col in enumerate(headers, 1):
            val = row[col]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = float(val) if isinstance(val, (float, np.floating)) else val
            if col == "Mediana ceny":
                cell.number_format = "#,##0.0000"
            cell.border = _thin_border()
            cell.font = Font(name="Arial", size=9)
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color=CLR_ALT_BG)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    _auto_width(ws)


def _build_settings_sheet(wb: Workbook, threshold_pct: float, file_name: str, df: pd.DataFrame):
    ws = wb.create_sheet("Ustawienia")
    ws.sheet_view.showGridLines = False

    data = [
        ("Parametr", "Wartość"),
        ("Próg odchylenia (%)", threshold_pct),
        ("Data wygenerowania", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Plik źródłowy", file_name),
        ("Liczba rekordów", len(df)),
        ("Liczba anomalii", int(df["Anomalia"].sum())),
        ("Liczba indeksów z anomaliami", int(df[df["Anomalia"]]["Index materiałowy"].nunique())),
        ("Udział anomalii (%)", round(df["Anomalia"].mean() * 100, 2)),
    ]

    for row_idx, (k, v) in enumerate(data, 1):
        kc = ws.cell(row=row_idx, column=1, value=k)
        vc = ws.cell(row=row_idx, column=2, value=v)
        if row_idx == 1:
            for c in (kc, vc):
                _header_style(c)
        else:
            kc.font = Font(name="Arial", bold=True, size=10)
            vc.font = Font(name="Arial", size=10)
        for c in (kc, vc):
            c.border = _thin_border()
            c.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
